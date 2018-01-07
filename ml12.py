import openpyxl

# book = openpyxl.load_workbook("stat_104102.xlsx")

# # print(book.get_sheet_names())

# sheet = book.worksheets[0]
# for row in sheet.rows:
#   for data in row:
#     print(data.value, end=" ")
#   print("", end="\n")

workbook = openpyxl.Workbook()
sheet = workbook.active

sheet["A1"] = "테스트 파일"
sheet["A2"] = "안녕하세요"
sheet.merge_cells("A1:C1")
sheet["A1"].font = openpyxl.styles.Font(size=20, color="FF0000")

workbook.save("newFile.xlsx")